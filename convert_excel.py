# -*- coding: utf-8 -*-
import openpyxl
import json
import os
import re
import sys
from datetime import datetime, timedelta

sys.stdout.reconfigure(encoding='utf-8')

desktop = os.path.join(os.environ['USERPROFILE'], 'Desktop')

# Use the new question bank file: 2026新题库0316（答案不确定标黄）.xlsx
source_file = '2026新题库0316（答案不确定标黄）.xlsx'
source_path = os.path.join(desktop, source_file)

if not os.path.exists(source_path):
    # Fallback: search for matching file
    files = [f for f in os.listdir(desktop) if '2026' in f and '0316' in f and f.endswith('.xlsx') and not f.startswith('~') and not f.startswith('副本')]
    if files:
        source_file = files[0]
        source_path = os.path.join(desktop, source_file)
    else:
        print("ERROR: Cannot find the question bank Excel file!")
        exit(1)

print(f"Loading: {source_file}")
wb = openpyxl.load_workbook(source_path)
ws = wb.active

# New file column layout:
# Col 1: 知识点 (Knowledge category)
# Col 2: 序号 (ID)
# Col 3: 题型 (Question type)
# Col 4: 题干 (Question text)
# Col 5: 选项A
# Col 6: 选项B
# Col 7: 选项C
# Col 8: 选项D
# Col 9: 答案 (Answer)
# Col 10: 标记 (Mark/Tag)
# Data starts from row 2 (row 1 is header)

def convert_cell_value(val):
    """Convert cell value to string, handling Excel date serial numbers."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        # Check if it looks like an Excel date serial number (common range)
        # Excel dates: 1 = 1900-1-1, typical range for years 1999-2026: ~36161-46388
        if 30000 <= val <= 50000 and val == int(val):
            try:
                # Excel date serial: day 1 = 1900-01-01, but Excel has a bug treating 1900 as leap year
                base_date = datetime(1899, 12, 30)
                date_val = base_date + timedelta(days=int(val))
                return date_val.strftime('%Y年%m月%d日')
            except Exception:
                pass
    # Handle datetime objects directly
    if isinstance(val, datetime):
        return val.strftime('%Y年%m月%d日')
    return str(val)

def is_yellow_fill(cell):
    """Check if a cell has a yellow background (indicating uncertain answer)."""
    try:
        fill = cell.fill
        if fill and fill.start_color and fill.start_color.rgb:
            color = str(fill.start_color.rgb).upper()
            # Common yellow colors in Excel
            if color in ('FFFFFF00', 'FFFFC000', 'FFFFEB9C', 'FFFFCC00', 'FFFFD700'):
                return True
            # Check for yellow-ish colors (high R, high G, low B)
            if len(color) == 8 and color != '00000000':
                try:
                    r = int(color[2:4], 16)
                    g = int(color[4:6], 16)
                    b = int(color[6:8], 16)
                    if r > 200 and g > 180 and b < 100:
                        return True
                except ValueError:
                    pass
    except Exception:
        pass
    return False

questions = []
uncertain_count = 0

for r in range(2, ws.max_row + 1):
    category = ws.cell(r, 1).value       # 知识点
    q_id = ws.cell(r, 2).value           # 序号
    q_type = ws.cell(r, 3).value         # 题型
    q_text = ws.cell(r, 4).value         # 题干
    opt_a = ws.cell(r, 5).value          # 选项A
    opt_b = ws.cell(r, 6).value          # 选项B
    opt_c = ws.cell(r, 7).value          # 选项C
    opt_d = ws.cell(r, 8).value          # 选项D
    answer = ws.cell(r, 9).value         # 答案
    mark = ws.cell(r, 10).value          # 标记

    if not q_text:
        continue

    # Check if answer is uncertain (yellow highlighted)
    answer_cell = ws.cell(r, 9)
    is_uncertain = is_yellow_fill(answer_cell)
    if is_uncertain:
        uncertain_count += 1

    # FIX: Use "is not None" instead of truthy check, so 0 values are preserved
    options = []
    if opt_a is not None:
        options.append({"key": "A", "text": convert_cell_value(opt_a)})
    if opt_b is not None:
        options.append({"key": "B", "text": convert_cell_value(opt_b)})
    if opt_c is not None:
        options.append({"key": "C", "text": convert_cell_value(opt_c)})
    if opt_d is not None:
        options.append({"key": "D", "text": convert_cell_value(opt_d)})

    # Determine type code
    q_type_str = str(q_type) if q_type else ""
    if '多' in q_type_str:
        type_code = 'multi'
    elif '判断' in q_type_str:
        type_code = 'judge'
    else:
        type_code = 'single'

    # Normalize answer format
    ans_str = str(answer).strip() if answer else ""

    # Handle judgment answers: 正确 -> A, 错误 -> B
    if type_code == 'judge':
        if '正确' in ans_str or ans_str == '对':
            ans_str = 'A'
        elif '错误' in ans_str or ans_str == '错':
            ans_str = 'B'

    # Clean up multi-choice answer separators
    # Remove all separators: commas, spaces, 、 etc. Keep only letters
    ans_clean = re.sub(r'[^A-Da-d]', '', ans_str)
    ans_clean = ans_clean.upper()

    # Sort letters for consistency
    if len(ans_clean) > 1:
        ans_clean = ''.join(sorted(set(ans_clean)))

    # Fallback: if cleaning removed everything, use original
    if not ans_clean:
        ans_clean = ans_str

    # Convert question text - handle potential datetime in question
    q_text_str = convert_cell_value(q_text) if q_text else ""
    
    # Convert category
    category_str = str(category).strip() if category else ""
    
    # Convert mark
    mark_str = str(mark).strip() if mark and str(mark) != 'None' else ""

    q_obj = {
        "id": int(q_id) if q_id else r - 1,
        "type": type_code,
        "typeName": q_type_str if q_type_str else "",
        "question": q_text_str,
        "options": options,
        "answer": ans_clean,
        "analysis": "",  # No analysis column in new file
        "category": category_str,
    }
    
    # Add uncertain flag if answer is highlighted yellow
    if is_uncertain:
        q_obj["uncertain"] = True
    
    # Add mark if present
    if mark_str:
        q_obj["mark"] = mark_str

    questions.append(q_obj)

wb.close()

# Write as JS
with open(r'd:\code_test\quiz-app\questions.js', 'w', encoding='utf-8') as f:
    f.write('const QUESTIONS = ')
    json.dump(questions, f, ensure_ascii=False, indent=None)
    f.write(';\n')

single_count = sum(1 for q in questions if q['type'] == 'single')
multi_count = sum(1 for q in questions if q['type'] == 'multi')
judge_count = sum(1 for q in questions if q['type'] == 'judge')

print(f"Converted {len(questions)} questions to questions.js")
print(f"Types: single={single_count}, multi={multi_count}, judge={judge_count}")
print(f"Uncertain answers (yellow): {uncertain_count}")

# Verify: check questions with option A = "0"
zero_a = [q for q in questions if any(o['key'] == 'A' and o['text'] == '0' for o in q['options'])]
print(f"Questions with A='0': {len(zero_a)}")
if zero_a:
    q = zero_a[0]
    print(f"  Sample: id={q['id']}, options={q['options']}, answer={q['answer']}")

# Verify: check for date-like values in options
date_opts = [q for q in questions if any('年' in o['text'] and '月' in o['text'] for o in q['options'])]
print(f"Questions with date options: {len(date_opts)}")
if date_opts:
    q = date_opts[0]
    print(f"  Sample: id={q['id']}, q={q['question'][:50]}, options={[o['text'] for o in q['options']]}")

# Verify: check judgment answers
judge_answers = {}
for q in questions:
    if q['type'] == 'judge':
        judge_answers[q['answer']] = judge_answers.get(q['answer'], 0) + 1
print(f"Judgment answers: {judge_answers}")

# Show some categories
categories = set(q.get('category', '') for q in questions if q.get('category'))
print(f"\nTotal knowledge categories: {len(categories)}")

# Show uncertain answer samples
uncertain_qs = [q for q in questions if q.get('uncertain')]
if uncertain_qs:
    print(f"\nUncertain answer samples (first 5):")
    for q in uncertain_qs[:5]:
        print(f"  id={q['id']}, type={q['type']}, answer={q['answer']}, q={q['question'][:60]}")
